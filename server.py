import os
import json
import sqlite3
import logging
import datetime
import io
import base64
import urllib.request
import urllib.parse
import openpyxl
import openpyxl.styles
import openpyxl.utils
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

import time

logger = logging.getLogger(__name__)

BIKE_SHEETS_CACHE = {"timestamp": 0, "data": {}}
TASKS_SHEETS_CACHE = {"timestamp": 0, "data": []}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if os.path.exists(os.path.join(BASE_DIR, "web_app", "index.html")):
    WEB_APP_DIR = os.path.join(BASE_DIR, "web_app")
else:
    WEB_APP_DIR = BASE_DIR

BIKES_DB_PATH = os.path.join(BASE_DIR, "bike_reports.db")
TASKS_DB_PATH = os.path.join(BASE_DIR, "tasks.db")
BOT_TOKEN = os.getenv("BOT_TOKEN", "8951006941:AAH2Wc2j2AH1aCvui1Bflr7puDStzHtwNNI").strip()
MASTER_APP_PASSWORD = os.getenv("MASTER_APP_PASSWORD", "7890").strip()

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_POSTGRES = True
except ImportError:
    HAS_POSTGRES = False

class PostgresRow:
    def __init__(self, data):
        self.data = data
        self.keys_list = list(data.keys())
    def __getitem__(self, key):
        if isinstance(key, int):
            return self.data[self.keys_list[key]]
        return self.data[key]
    def keys(self):
        return self.data.keys()
    def get(self, key, default=None):
        return self.data.get(key, default)
    def __contains__(self, key):
        return key in self.data

class PostgresCursorWrapper:
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, query, params=None):
        query = query.replace("?", "%s")
        if "INSERT OR IGNORE" in query:
            if "rich_cities" in query:
                query = query.replace("INSERT OR IGNORE INTO rich_cities", "INSERT INTO rich_cities") + " ON CONFLICT (name) DO NOTHING"
            elif "managed_bots" in query:
                query = query.replace("INSERT OR IGNORE", "INSERT")
        query = query.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
        query = query.replace("AUTOINCREMENT", "")
        query = query.replace("REAL DEFAULT 0", "DOUBLE PRECISION DEFAULT 0")
        if "LIKE '%' ||" in query:
            query = query.replace("LIKE '%' ||", "ILIKE '%' ||")
            
        if params is not None:
            self._cursor.execute(query, params)
        else:
            self._cursor.execute(query)
            
    def executemany(self, query, seq):
        query = query.replace("?", "%s")
        if "INSERT OR IGNORE" in query:
            if "rich_cities" in query:
                query = query.replace("INSERT OR IGNORE INTO rich_cities", "INSERT INTO rich_cities") + " ON CONFLICT (name) DO NOTHING"
            else:
                query = query.replace("INSERT OR IGNORE", "INSERT")
        self._cursor.executemany(query, seq)
        
    def fetchone(self):
        res = self._cursor.fetchone()
        return PostgresRow(res) if res else None
        
    def fetchall(self):
        res = self._cursor.fetchall()
        return [PostgresRow(r) for r in res] if res else []
        
    @property
    def lastrowid(self):
        try:
            self._cursor.execute("SELECT lastval() AS id")
            row = self._cursor.fetchone()
            return row["id"] if row else None
        except Exception:
            return None

class PostgresConnectionWrapper:
    def __init__(self, conn):
        self._conn = conn
        
    def cursor(self):
        return PostgresCursorWrapper(self._conn.cursor())
        
    def commit(self):
        self._conn.commit()
        
    def close(self):
        self._conn.close()
        
    @property
    def row_factory(self):
        return None
        
    @row_factory.setter
    def row_factory(self, val):
        pass

_orig_sqlite3_connect = sqlite3.connect

def connect_wrapper(db_path, *args, **kwargs):
    db_url = os.getenv("DATABASE_URL")
    if db_url and HAS_POSTGRES:
        conn = psycopg2.connect(db_url)
        return PostgresConnectionWrapper(conn)
    else:
        return _orig_sqlite3_connect(db_path, *args, **kwargs)

sqlite3.connect = connect_wrapper

# Daily rates by role (UZS) for Yandex Eats project
ROLE_DAILY_RATES = {
    "Тимлиды": 300000,
    "Активаторы": 250000,
    "Регион": 220000,
    "Ресепшн": 210000,
    "Склад": 240000,
    "Поддержка": 200000,
    "ТМ": 350000,
    "Лайвопс": 280000,
    "Клининг": 180000,
    "Лавка": 220000,
    "Общий": 230000
}

# Cache for latest parsed schedule result
LATEST_EXCEL_PARSED = {
    "employees": [],
    "summary": {"total_employees": 0, "total_advance": 0, "total_net": 0, "total_tax": 0, "total_fot": 0}
}

def init_tables():
    if not os.getenv("DATABASE_URL") and not os.path.exists(BIKES_DB_PATH):
        return
    try:
        conn = sqlite3.connect(BIKES_DB_PATH)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS rich_cities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                total_bikes INTEGER DEFAULT 50,
                created_at TEXT
            )
        """)
        c.execute("SELECT COUNT(*) FROM rich_cities")
        if c.fetchone()[0] == 0:
            c.executemany("INSERT INTO rich_cities (name, total_bikes) VALUES (?, ?)", [
                ("Ташкент (Rich)", 100),
                ("Самарканд (Rich)", 50),
                ("Бухара (Rich)", 30),
            ])

        c.execute("""
            CREATE TABLE IF NOT EXISTS rich_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                city TEXT,
                report_date TEXT,
                issued INTEGER DEFAULT 0,
                returned INTEGER DEFAULT 0,
                comment TEXT,
                status TEXT DEFAULT 'Active'
            )
        """)
        try:
            c.execute("ALTER TABLE rich_reports ADD COLUMN comment TEXT")
        except Exception:
            pass
        c.execute("""
            CREATE TABLE IF NOT EXISTS payroll (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_name TEXT,
                role TEXT DEFAULT 'Сотрудник',
                exact_shifts REAL DEFAULT 0,
                hours INTEGER DEFAULT 0,
                advance_amount REAL DEFAULT 0,
                salary_amount REAL DEFAULT 0,
                tax_amount REAL DEFAULT 0,
                total_fot REAL DEFAULT 0,
                payment_date TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS managed_bots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bot_name TEXT NOT NULL,
                bot_token TEXT NOT NULL,
                bot_username TEXT DEFAULT '',
                project_type TEXT DEFAULT 'FlitGo',
                city_name TEXT DEFAULT 'Ташкент',
                report_type TEXT DEFAULT 'Отчётность',
                is_active INTEGER DEFAULT 1,
                created_at TEXT
            )
        """)
        try:
            c.execute("ALTER TABLE managed_bots ADD COLUMN bot_username TEXT DEFAULT ''")
        except Exception:
            pass
        c.execute("SELECT COUNT(*) FROM rich_cities")
        if c.fetchone()[0] == 0:
            c.executemany("INSERT INTO rich_cities (name, total_bikes) VALUES (?, ?)", [
                ("Ташкент (Rich)", 100),
                ("Самарканд (Rich)", 50),
                ("Бухара (Rich)", 30),
            ])

        c.execute("SELECT COUNT(*) FROM managed_bots")
        if c.fetchone()[0] == 0:
            now_str = datetime.datetime.now().strftime("%d.%m.%Y")
            c.executemany("INSERT INTO managed_bots (bot_name, bot_token, project_type, city_name, report_type, is_active, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)", [
                ("FlitGo Bike Report Bot", "8123456789:AAXXXXXXXXXXXXXX", "FlitGo", "Ташкент", "Отчёт по байкам", 1, now_str),
                ("FlitGo Task Bot", "8987654321:AAYYYYYYYYYYYYYY", "FlitGo", "Все города", "Управление задачами", 1, now_str),
                ("Rich Hybrid Bot", "8951006941:AAH2Wc2j2AH1aCvui1Bflr7puDStzHtwNNI", "Rich", "Все города", "Гибриды Rich", 1, now_str),
            ])
        c.execute("""
            CREATE TABLE IF NOT EXISTS task_archive (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_label TEXT NOT NULL,
                period_start TEXT NOT NULL,
                period_end TEXT NOT NULL,
                assignee TEXT NOT NULL,
                total_tasks INTEGER DEFAULT 0,
                done_tasks INTEGER DEFAULT 0,
                avg_rating REAL DEFAULT 0,
                efficiency INTEGER DEFAULT 0,
                archived_at TEXT
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Failed to init tables: {e}")

init_tables()

import re

def parse_excel_schedule(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    for sname in wb.sheetnames:
        if "июль" in sname.lower() or "2026" in sname:
            sheet = wb[sname]
            break

    current_role = "Общий"
    parsed_employees = []
    total_advance = 0
    total_net = 0
    total_tax = 0
    total_fot = 0

    for r in range(1, sheet.max_row + 1):
        # Col A = FIO
        fio_val = sheet.cell(row=r, column=1).value
        if not fio_val:
            fio_val = sheet.cell(row=r, column=2).value
        
        if not fio_val:
            continue
            
        fio_str = str(fio_val).strip()
        if not fio_str:
            continue

        # Check for role headers
        matched_role = None
        for r_name in ROLE_DAILY_RATES.keys():
            if r_name.lower() in fio_str.lower():
                matched_role = r_name
                break
        if matched_role:
            current_role = matched_role
            if len(fio_str.split()) <= 3:
                continue

        # Filter out non-names
        if not re.search(r'[a-zA-Zа-яА-ЯёЁ]{2,}', fio_str) or any(k in fio_str.lower() for k in ["фио", "итого", "всего", "смены"]):
            continue

        # Col I = [Оклад]
        salary_val = 0.0
        try:
            salary_val = float(sheet.cell(row=r, column=9).value or 0) # Col I
        except Exception:
            salary_val = 0.0

        if salary_val == 0.0:
            for c_idx in [8, 10, 7]:
                try:
                    v_sal = float(sheet.cell(row=r, column=c_idx).value or 0)
                    if v_sal > 500000:
                        salary_val = v_sal
                        break
                except Exception:
                    pass

        if salary_val == 0.0:
            daily_rate = ROLE_DAILY_RATES.get(current_role, 230000)
            salary_val = daily_rate * 15.0

        # Col C = [Смен_факт]
        try:
            exact_shifts = float(sheet.cell(row=r, column=3).value or 0) # Col C
        except Exception:
            exact_shifts = 0.0

        if exact_shifts == 0.0:
            try:
                exact_shifts = float(sheet.cell(row=r, column=4).value or 0) # Col D
            except Exception:
                exact_shifts = 15.0

        # Col E = [Дни_отпуска]
        try:
            vacation_days = int(sheet.cell(row=r, column=5).value or 0) # Col E
        except Exception:
            vacation_days = 0

        # Col F = [Дни_больничного]
        try:
            sick_days = int(sheet.cell(row=r, column=6).value or 0) # Col F
        except Exception:
            sick_days = 0

        if vacation_days == 0 and sick_days == 0:
            for col_idx in range(12, min(48, sheet.max_column + 1)):
                val = sheet.cell(row=r, column=col_idx).value
                if val is not None:
                    val_s = str(val).strip().upper()
                    if val_s == 'Б':
                        sick_days += 1
                    elif val_s == 'О':
                        vacation_days += 1

        # User specified formulas:
        # 1. Стоимость 1 смены = [Оклад] / 15
        shift_rate = salary_val / 15.0

        # 2. Оплата за отработанные смены (основные + доп. смены) = [Смен_факт] * Стоимость 1 смены
        shifts_pay = exact_shifts * shift_rate

        # 3. Оплата отпуска = ([Оклад] / 25.3) * [Дни_отпуска]
        vacation_pay = (salary_val / 25.3) * vacation_days

        # 4. Оплата больничного = (Стоимость 1 смены * 0.6) * [Дни_больничного] (60%)
        sick_pay = (shift_rate * 0.60) * sick_days

        # Общее начисление за месяц (Начислено чистыми)
        total_earned_net = round(shifts_pay + vacation_pay + sick_pay)

        # Аванс = 50% от оклада (уже выплачено в середине месяца)
        advance = round(salary_val * 0.50)

        # Зарплата на руки к выдаче в конце месяца = Общая ЗП - Аванс
        net_pay = max(0, total_earned_net - advance)

        # Всего выплачено за месяц (Аванс + ЗП)
        total_payout = advance + net_pay

        # Налог (ФОТ 24% ТК РУз)
        tax = round((total_payout / 0.88) * 0.24) if total_payout > 0 else 0
        fot = total_payout + tax

        total_advance += advance
        total_net += net_pay
        total_tax += tax
        total_fot += fot

        parsed_employees.append({
            "name": fio_str,
            "role": current_role,
            "exact_shifts": round(exact_shifts, 1),
            "base_salary": round(salary_val),
            "sick_days": sick_days,
            "vacation_days": vacation_days,
            "advance": advance,
            "net_pay": net_pay,
            "tax": tax,
            "fot": fot
        })

    result = {
        "employees": parsed_employees,
        "summary": {
            "total_employees": len(parsed_employees),
            "total_advance": total_advance,
            "total_net": total_net,
            "total_tax": total_tax,
            "total_fot": total_fot
        }
    }
    LATEST_EXCEL_PARSED["employees"] = parsed_employees
    LATEST_EXCEL_PARSED["summary"] = result["summary"]
    return result

def generate_payroll_excel_bytes(employees, summary):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Итоговая Ведомость ЗП"

    header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = openpyxl.styles.Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = openpyxl.styles.Font(name="Calibri", size=14, bold=True, color="1E293B")

    ws.append(["Итоговая Расчётная Ведомость (ФИО - Аванс - ЗП - ФОТ/Налог)"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([])

    headers = ["№", "ФИО сотрудника", "Отдел / Роль", "Смены", "Больничные (Б)", "Отпуск (О)", "Прогулы (П)", "Аванс (сум)", "ЗП на руки (сум)", "ФОТ (Налог 24% ТК РУз)", "Всего расходы (сум)"]
    ws.append(headers)

    header_row = 3
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    row_idx = 4
    for idx, emp in enumerate(employees, start=1):
        ws.append([
            idx,
            emp.get("name", ""),
            emp.get("role", "Сотрудник"),
            emp.get("exact_shifts", 0),
            emp.get("sick_days", 0),
            emp.get("vacation_days", 0),
            emp.get("absent_days", 0),
            emp.get("advance", 0),
            emp.get("net_pay", 0),
            emp.get("tax", 0),
            emp.get("fot", 0)
        ])
        row_idx += 1

    ws.append([
        "ИТОГО",
        "",
        "",
        "",
        "",
        "",
        "",
        summary.get("total_advance", 0),
        summary.get("total_net", 0),
        summary.get("total_tax", 0),
        summary.get("total_fot", 0)
    ])

    tot_row = row_idx
    ws.cell(row=tot_row, column=1).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=tot_row, column=8).font = openpyxl.styles.Font(bold=True, color="059669")
    ws.cell(row=tot_row, column=9).font = openpyxl.styles.Font(bold=True, color="0284C7")
    ws.cell(row=tot_row, column=10).font = openpyxl.styles.Font(bold=True, color="D97706")
    ws.cell(row=tot_row, column=11).font = openpyxl.styles.Font(bold=True, color="6D28D9")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def send_telegram_file(chat_id, file_bytes, filename="Payroll_Report_TK_RUZ.xlsx"):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
    boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
    
    body = []
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="chat_id"'.encode())
    body.append(b"")
    body.append(str(chat_id).encode())
    
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="caption"'.encode())
    body.append(b"")
    body.append("📊 Итоговая расчётная ведомость по ТК РУз (Yandex Eats)".encode("utf-8"))
    
    body.append(f"--{boundary}".encode())
    body.append(f'Content-Disposition: form-data; name="document"; filename="{filename}"'.encode())
    body.append(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    body.append(b"")
    body.append(file_bytes)
    
    body.append(f"--{boundary}--\r\n".encode())
    
    payload_data = b"\r\n".join(body)
    
    req = urllib.request.Request(
        url,
        data=payload_data,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}
    )
    resp = urllib.request.urlopen(req)
    return json.loads(resp.read().decode("utf-8"))

class MasterHubHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=WEB_APP_DIR, **kwargs)

    def end_headers(self):
        self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_GET(self):
        if self.path in ("/", "/health", "/ping"):
            index_path = os.path.join(WEB_APP_DIR, "index.html")
            if not os.path.exists(index_path):
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(b"OK - Service is running")
                return
        if self.path.startswith("/api/"):
            self.handle_api_get()
        else:
            super().do_GET()

    def do_POST(self):
        if self.path.startswith("/api/"):
            self.handle_api_post()
        else:
            self.send_error(404, "Not Found")

    def handle_api_get(self):
        path = self.path.split("?")[0]
        if path == "/api/dashboard":
            self.send_json_response(self.get_dashboard_data())
        elif path == "/api/cities":
            self.send_json_response(self.get_cities_data())
        elif path == "/api/reports":
            self.send_json_response(self.get_reports_data())
        elif path == "/api/broken_bikes_by_city":
            self.send_json_response(self.get_broken_bikes_by_city())
        elif path == "/api/tasks":
            self.send_json_response(self.get_tasks_data())
        elif path == "/api/team_leads_tasks":
            params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            month = params.get("month", [None])[0]
            self.send_json_response(self.get_team_leads_task_stats(month))
        elif path == "/api/users":
            self.send_json_response(self.get_users_data())
        elif path == "/api/rich/cities":
            self.send_json_response(self.get_rich_cities())
        elif path == "/api/rich/reports":
            self.send_json_response(self.get_rich_reports())
        elif path == "/api/rich/stats":
            self.send_json_response(self.get_rich_stats())
        elif path == "/api/payroll":
            self.send_json_response(self.get_payroll_data())
        elif path == "/api/payroll/download_excel":
            excel_bytes = generate_payroll_excel_bytes(LATEST_EXCEL_PARSED["employees"], LATEST_EXCEL_PARSED["summary"])
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="Payroll_Report_TK_RUZ.xlsx"')
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(excel_bytes)
        elif path == "/api/bots":
            self.send_json_response(self.get_bots_data())
        elif path == "/api/tasks/archives":
            self.send_json_response(self.get_archives())
        elif path == "/api/tasks/dynamics":
            params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            date_from = params.get("date_from", [None])[0]
            date_to = params.get("date_to", [None])[0]
            assignee = params.get("assignee", [None])[0]
            self.send_json_response(self.get_tasks_dynamics(date_from, date_to, assignee))
        else:
            self.send_error(404, "API Endpoint Not Found")

    def handle_api_post(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length)
        try:
            payload = json.loads(body.decode("utf-8")) if body else {}
        except Exception:
            payload = {}

        path = self.path.split("?")[0]
        if path == "/api/auth/verify":
            password = str(payload.get("password", "")).strip()
            if password == MASTER_APP_PASSWORD:
                self.send_json_response({"success": True, "message": "Авторизация успешна"})
            else:
                self.send_json_response({"success": False, "error": "Неверный пароль доступа"}, status=401)
            return

        if path in ("/api/cities/update", "/api/cities/update_total"):
            city_id = payload.get("city_id")
            total_bikes = payload.get("total_bikes")
            if city_id and total_bikes:
                self.update_city_total(city_id, total_bikes)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Invalid params"}, status=400)
        elif path == "/api/users/toggle_access":
            user_id = payload.get("user_id")
            is_active = payload.get("is_active", 1)
            if user_id is not None:
                self.toggle_user_access(user_id, is_active)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Invalid params"}, status=400)
        elif path == "/api/tasks/create":
            task_text = payload.get("task_text", "").strip()
            assignee = payload.get("assignee", "Сотрудник").strip()
            sla_deadline = payload.get("sla_deadline", "24 часа").strip()
            priority = payload.get("priority", "Medium").strip()
            city = payload.get("city", "Ташкент").strip()
            if task_text:
                self.create_task(task_text, assignee, sla_deadline, priority, city)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Task text is required"}, status=400)
        elif path == "/api/tasks/complete":
            task_id = payload.get("task_id")
            if task_id:
                self.complete_task(task_id)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Task ID required"}, status=400)
        elif path == "/api/tasks/rate":
            task_id = payload.get("task_id")
            rating = payload.get("rating", 5)
            comment = payload.get("rating_comment", "")
            if task_id:
                self.rate_task(task_id, rating, comment)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Task ID required"}, status=400)
        elif path == "/api/rich/cities/add":
            name = payload.get("name", "").strip()
            total_bikes = payload.get("total_bikes", 50)
            if name:
                self.add_rich_city(name, total_bikes)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "City name required"}, status=400)
        elif path == "/api/rich/cities/update":
            city_id = payload.get("city_id")
            total_bikes = payload.get("total_bikes")
            if city_id and total_bikes:
                self.update_rich_city(city_id, total_bikes)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Invalid params"}, status=400)
        elif path == "/api/payroll/calculate":
            advance = float(payload.get("advance", 0))
            salary = float(payload.get("salary", 0))
            total_pay = advance + salary
            tax = round((total_pay / 0.88) * 0.24) if total_pay > 0 else 0
            fot = total_pay + tax
            self.send_json_response({
                "advance": advance,
                "salary": salary,
                "total_net": total_pay,
                "tax": tax,
                "total_fot": fot
            })
        elif path == "/api/payroll/add":
            emp_name = payload.get("employee_name", "Сотрудник").strip()
            advance = float(payload.get("advance", 0))
            salary = float(payload.get("salary", 0))
            total_pay = advance + salary
            tax = round((total_pay / 0.88) * 0.24) if total_pay > 0 else 0
            fot = total_pay + tax
            self.add_payroll_entry(emp_name, advance, salary, tax, fot)
            self.send_json_response({"status": "ok"})
        elif path == "/api/payroll/upload_schedule":
            file_base64 = payload.get("file_b64", "")
            if file_base64:
                try:
                    file_bytes = base64.b64decode(file_base64)
                    result = parse_excel_schedule(file_bytes)
                    self.send_json_response(result)
                except Exception as e:
                    logger.error(f"Failed to parse schedule excel: {e}")
                    self.send_json_response({"error": f"Error parsing excel: {str(e)}"}, status=400)
            else:
                self.send_json_response({"error": "file_b64 is required"}, status=400)
        elif path == "/api/payroll/send_telegram":
            chat_id = payload.get("chat_id")
            if not chat_id:
                chat_id = "560410710"
            try:
                excel_bytes = generate_payroll_excel_bytes(LATEST_EXCEL_PARSED["employees"], LATEST_EXCEL_PARSED["summary"])
                res_tg = send_telegram_file(chat_id, excel_bytes)
                self.send_json_response({"status": "ok", "tg_response": res_tg})
            except Exception as e:
                logger.error(f"Error sending telegram file: {e}")
                self.send_json_response({"error": str(e)}, status=400)
        elif path == "/api/bots/add":
            bot_name = payload.get("bot_name", "").strip()
            bot_token = payload.get("bot_token", "").strip()
            bot_username = payload.get("bot_username", "").strip().lstrip("@")
            project_type = payload.get("project_type", "FlitGo").strip()
            city_name = payload.get("city_name", "Ташкент").strip()
            report_type = payload.get("report_type", "Отчётность").strip()
            if bot_name and bot_token:
                self.add_managed_bot(bot_name, bot_token, bot_username, project_type, city_name, report_type)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Bot name and token are required"}, status=400)
        elif path == "/api/bots/toggle":
            bot_id = payload.get("bot_id")
            is_active = payload.get("is_active", 1)
            if bot_id:
                self.toggle_managed_bot(bot_id, is_active)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Bot ID required"}, status=400)
        elif path == "/api/bots/update":
            bot_id = payload.get("bot_id")
            bot_name = payload.get("bot_name", "").strip()
            bot_token = payload.get("bot_token", "").strip()
            bot_username = payload.get("bot_username", "").strip().lstrip("@")
            project_type = payload.get("project_type", "FlitGo").strip()
            city_name = payload.get("city_name", "Ташкент").strip()
            report_type = payload.get("report_type", "Отчётность").strip()
            if bot_id and bot_name:
                self.update_managed_bot(bot_id, bot_name, bot_token, bot_username, project_type, city_name, report_type)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Bot ID and name required"}, status=400)
        elif path == "/api/bots/delete":
            bot_id = payload.get("bot_id")
            if bot_id:
                self.delete_managed_bot(bot_id)
                self.send_json_response({"status": "ok"})
            else:
                self.send_json_response({"error": "Bot ID required"}, status=400)
        elif path == "/api/tasks/archive":
            period_label = payload.get("period_label", "").strip()
            period_start = payload.get("period_start", "").strip()
            period_end = payload.get("period_end", "").strip()
            if period_label and period_start and period_end:
                result = self.archive_current_period(period_label, period_start, period_end)
                self.send_json_response(result)
            else:
                self.send_json_response({"error": "period_label, period_start, period_end are required"}, status=400)
        else:
            self.send_error(404, "API Endpoint Not Found")

    def send_json_response(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def get_dashboard_data(self):
        cities_list = self.get_cities_data()
        tot_bikes = sum(int(c.get("total_bikes") or 0) for c in cities_list)
        tot_issued = sum(int(c.get("issued") or 0) for c in cities_list)
        share_on_line = round((tot_issued / tot_bikes) * 100) if tot_bikes > 0 else 0

        tot_tasks = len(self.get_tasks_data())

        return {
            "total_bikes": tot_bikes,
            "share_on_line": share_on_line,
            "active_tasks": tot_tasks,
            "total_users": len(cities_list),
        }

    def get_cities_data(self):
        default_cities_fallback = [
            {"id": 1, "name": "Ташкент", "total_bikes": 1670, "has_bike_types": 0},
            {"id": 2, "name": "Самарканд", "total_bikes": 200, "has_bike_types": 0},
            {"id": 3, "name": "Фергана", "total_bikes": 80, "has_bike_types": 0},
            {"id": 4, "name": "Андижан", "total_bikes": 50, "has_bike_types": 0},
            {"id": 5, "name": "Бухара", "total_bikes": 30, "has_bike_types": 0},
            {"id": 6, "name": "Навои", "total_bikes": 30, "has_bike_types": 0},
            {"id": 7, "name": "Карши", "total_bikes": 30, "has_bike_types": 0},
            {"id": 8, "name": "Ургенч", "total_bikes": 30, "has_bike_types": 0},
            {"id": 9, "name": "Нукус", "total_bikes": 30, "has_bike_types": 0},
            {"id": 10, "name": "Коканд", "total_bikes": 25, "has_bike_types": 0},
            {"id": 11, "name": "Наманган", "total_bikes": 25, "has_bike_types": 0},
        ]
        
        raw_rows = []
        try:
            if not os.path.exists(BIKES_DB_PATH):
                init_local_master_dbs()
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT c.id, c.name, c.total_bikes, c.has_bike_types FROM cities c ORDER BY CASE WHEN c.name LIKE '%Ташкент%' THEN 1 ELSE 2 END, c.id ASC")
            raw_rows = [dict(r) for r in c.fetchall()]
            conn.close()
        except Exception as e:
            logger.error(f"Error querying cities from db: {e}")

        if not raw_rows:
            raw_rows = default_cities_fallback

        # Live Google Sheets fallback for bike reports (with 15s in-memory cache to prevent 429 Quota Exceeded)
        now_time = time.time()
        sheet_reports = {}
        if now_time - BIKE_SHEETS_CACHE["timestamp"] < 15 and BIKE_SHEETS_CACHE["data"]:
            sheet_reports = BIKE_SHEETS_CACHE["data"]
        else:
            creds_json_env = os.getenv("GOOGLE_CREDENTIALS_JSON")
            if creds_json_env:
                try:
                    import gspread
                    from google.oauth2.service_account import Credentials
                    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                    s_clean = creds_json_env.strip().strip("'").strip('"')
                    info = json.loads(s_clean)
                    if isinstance(info.get("private_key"), str):
                        info["private_key"] = info["private_key"].replace("\\n", "\n")
                    creds = Credentials.from_service_account_info(info, scopes=scopes)
                    client = gspread.authorize(creds)
                    spreadsheet = client.open_by_key("1Oskxt5oHfO50PDn47I_7rbn4KGfEoy_JcVsn3mBIiyw")
                    for ws in spreadsheet.worksheets():
                        title = ws.title
                        city_name = title.replace("Байки", "").strip() if "Байки" in title else title.strip()
                        if not city_name:
                            continue
                        rows = ws.get_all_values()
                        if len(rows) > 1:
                            latest = rows[-1]
                            is_new_format = (len(latest) >= 12 and not str(latest[0]).strip().startswith("202"))

                            if is_new_format:
                                issued_val = latest[5] if len(latest) > 5 else "0"
                                broken_val = latest[8] if len(latest) > 8 else "0"
                                date_val = latest[2] if len(latest) > 2 else ""
                            else:
                                headers = [str(h).strip() for h in rows[0]]
                                iss_idx = headers.index("В поездке") if "В поездке" in headers else (headers.index("Всего на линии") if "Всего на линии" in headers else 4)
                                brok_idx = headers.index("Сломанные") if "Сломанные" in headers else (headers.index("Сломанные байки") if "Сломанные байки" in headers else 6)
                                date_idx = headers.index("Дата отчета") if "Дата отчета" in headers else (headers.index("Дата") if "Дата" in headers else 1)

                                issued_val = latest[iss_idx] if len(latest) > iss_idx else "0"
                                broken_val = latest[brok_idx] if len(latest) > brok_idx else "0"
                                date_val = latest[date_idx] if len(latest) > date_idx else ""

                            try:
                                iss_num = int(issued_val)
                            except (ValueError, TypeError):
                                iss_num = 0
                            try:
                                brok_num = int(broken_val)
                            except (ValueError, TypeError):
                                brok_num = 0

                            sheet_reports[city_name.lower()] = {
                                "issued": iss_num,
                                "broken": brok_num,
                                "report_date": date_val
                            }
                    BIKE_SHEETS_CACHE["data"] = sheet_reports
                    BIKE_SHEETS_CACHE["timestamp"] = now_time
                except Exception as e:
                    logger.error(f"Failed to fetch bike reports from Google Sheets: {e}")
                    if BIKE_SHEETS_CACHE["data"]:
                        sheet_reports = BIKE_SHEETS_CACHE["data"]

        result = []
        for r in raw_rows:
            tot = int(r.get("total_bikes") or 0)
            c_name_lower = r["name"].lower()

            iss = 0
            broken = 0
            r_date = r.get("report_date")

            try:
                iss = int(r.get("issued") or 0)
                broken = int(r.get("broken_bikes") or 0)
            except (ValueError, TypeError):
                pass

            # Always prioritize live Google Sheets report data if available
            for sheet_city, rep_data in sheet_reports.items():
                if sheet_city in c_name_lower or c_name_lower in sheet_city:
                    try:
                        iss = int(rep_data["issued"])
                    except Exception:
                        pass
                    try:
                        broken = int(rep_data["broken"])
                    except Exception:
                        pass
                    r_date = rep_data["report_date"]
                    break

            pct = round((iss / tot) * 100) if tot > 0 else 0
            pct = min(pct, 100)

            result.append({
                "id": r["id"],
                "name": r["name"],
                "total_bikes": tot,
                "issued": iss,
                "percent_online": pct,
                "broken_bikes": broken,
                "report_date": r_date
            })
        return result

    def update_city_total(self, city_id: int, total_bikes: int):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            c.execute("UPDATE cities SET total_bikes = ? WHERE id = ?", (total_bikes, city_id))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to update city total: {e}")

    def get_reports_data(self):
        if not os.path.exists(BIKES_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT id, username, city, report_date, issued, returned, total_in_trip, broken_bikes FROM bike_reports ORDER BY id DESC LIMIT 5")
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return rows
        except Exception as e:
            logger.error(f"Failed to get reports: {e}")
            return []

    def get_broken_bikes_by_city(self):
        cities = self.get_cities_data()
        res = []
        for c in cities:
            try:
                broken = int(c.get("broken_bikes") or 0)
            except Exception:
                broken = 0
            if broken > 0:
                res.append({
                    "city": c.get("name", ""),
                    "broken_bikes": broken,
                    "report_date": c.get("report_date", ""),
                    "username": "Партнёр"
                })
        return res

    def get_tasks_data(self):
        if os.path.exists(TASKS_DB_PATH):
            try:
                conn = sqlite3.connect(TASKS_DB_PATH)
                conn.row_factory = sqlite3.Row
                c = conn.cursor()
                c.execute("SELECT id, task_text, assignee, author, sla_deadline, created_at, status, COALESCE(priority, 'Medium') as priority, COALESCE(city, 'Ташкент') as city, COALESCE(rating, 0) as rating, COALESCE(rating_comment, '') as rating_comment FROM tasks ORDER BY id DESC LIMIT 50")
                rows = [dict(r) for r in c.fetchall()]
                conn.close()
                if rows:
                    return rows
            except Exception as e:
                logger.error(f"Failed to get tasks from sqlite: {e}")

        # Fallback: Read live tasks from Google Sheets (with 15s cache to prevent 429 Quota Exceeded)
        now_time = time.time()
        if now_time - TASKS_SHEETS_CACHE["timestamp"] < 15 and TASKS_SHEETS_CACHE["data"]:
            return TASKS_SHEETS_CACHE["data"]

        creds_json_env = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if creds_json_env:
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                s_clean = creds_json_env.strip().strip("'").strip('"')
                info = json.loads(s_clean)
                if isinstance(info.get("private_key"), str):
                    info["private_key"] = info["private_key"].replace("\\n", "\n")
                creds = Credentials.from_service_account_info(info, scopes=scopes)
                client = gspread.authorize(creds)
                spreadsheet = client.open_by_key("14lJVvDmK9LOAERAo9twp3Ak-FEdvlrzu-8FywP2dTn4")
                sheet = spreadsheet.sheet1
                rows = sheet.get_all_values()
                tasks = []
                if len(rows) > 1:
                    headers = [str(h).strip() for h in rows[0]]
                    id_idx = headers.index("ID Задачи") if "ID Задачи" in headers else 0
                    text_idx = headers.index("Текст задачи") if "Текст задачи" in headers else 1
                    ass_idx = headers.index("Исполнитель") if "Исполнитель" in headers else 2
                    aut_idx = headers.index("Постановщик") if "Постановщик" in headers else 3
                    sla_idx = headers.index("Срок / SLA") if "Срок / SLA" in headers else 4
                    date_idx = headers.index("Дата создания") if "Дата создания" in headers else 5
                    stat_idx = headers.index("Статус") if "Статус" in headers else 6
                    init_rat_idx = headers.index("Первоначальная оценка") if "Первоначальная оценка" in headers else (headers.index("Оценка") if "Оценка" in headers else 7)
                    disp_idx = headers.index("Причина оспаривания") if "Причина оспаривания" in headers else (headers.index("Комментарий / Оспаривание") if "Комментарий / Оспаривание" in headers else 8)
                    final_rat_idx = headers.index("Последняя оценка") if "Последняя оценка" in headers else 9

                    for i, r in enumerate(rows[1:], start=1):
                        init_rat = r[init_rat_idx] if len(r) > init_rat_idx else "0"
                        disp_val = r[disp_idx] if len(r) > disp_idx else ""
                        final_rat = r[final_rat_idx] if len(r) > final_rat_idx else ""

                        try:
                            init_num = int(str(init_rat).replace("/5", "").strip())
                        except Exception:
                            init_num = 0

                        try:
                            final_num = int(str(final_rat).replace("/5", "").strip())
                        except Exception:
                            final_num = 0

                        is_disputed = bool(disp_val.strip() and not final_rat.strip())

                        tasks.append({
                            "id": r[id_idx] if len(r) > id_idx and r[id_idx] else i,
                            "task_text": r[text_idx] if len(r) > text_idx else "",
                            "assignee": r[ass_idx] if len(r) > ass_idx else "",
                            "author": r[aut_idx] if len(r) > aut_idx else "",
                            "sla_deadline": r[sla_idx] if len(r) > sla_idx else "",
                            "created_at": r[date_idx] if len(r) > date_idx else "",
                            "status": r[stat_idx] if len(r) > stat_idx else "Active",
                            "priority": "Medium",
                            "city": "Ташкент",
                            "rating": final_num if final_num > 0 else (init_num if not is_disputed else 0),
                            "initial_rating": init_num,
                            "final_rating": final_num,
                            "is_disputed": is_disputed,
                            "rating_comment": disp_val
                        })
                res_tasks = tasks[::-1]
                TASKS_SHEETS_CACHE["data"] = res_tasks
                TASKS_SHEETS_CACHE["timestamp"] = now_time
                return res_tasks
            except Exception as e:
                logger.error(f"Failed to fetch tasks from Google Sheets: {e}")
                if TASKS_SHEETS_CACHE["data"]:
                    return TASKS_SHEETS_CACHE["data"]

        return []

    def create_task(self, task_text: str, assignee: str, sla_deadline: str, priority: str = "Medium", city: str = "Ташкент"):
        if not os.path.exists(TASKS_DB_PATH):
            return
        try:
            conn = sqlite3.connect(TASKS_DB_PATH)
            c = conn.cursor()
            now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
            c.execute(
                "INSERT INTO tasks (task_text, assignee, author, sla_deadline, created_at, status, priority, city) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (task_text, assignee, "Руководитель", sla_deadline, now_str, "Active", priority, city)
            )
            task_id = c.lastrowid
            conn.commit()
            conn.close()

            bot_token = "8666306951:AAEJ9z2F0t4I2mj2IMPE8TygL6a2k_5ob6g"
            chat_id = "-1002638798110"

            prio_str = "🔴 Срочно (High)" if priority == "High" else ("🟡 Средний" if priority == "Medium" else "🟢 Обычный")

            msg_text = (
                f"📌 <b>НОВАЯ ЗАДАЧА #{task_id}</b>\n\n"
                f"📋 <b>Описание:</b> {task_text}\n"
                f"👤 <b>Исполнитель:</b> {assignee}\n"
                f"⚡️ <b>Приоритет:</b> {prio_str}\n"
                f"⏱ <b>SLA / Срок:</b> {sla_deadline}\n"
                f"✍️ <b>Постановщик:</b> Руководитель (@orzmkh)\n\n"
                f"🌐 <i>Поставлена через Центр Управления Master Hub</i>"
            )

            payload = {
                "chat_id": chat_id,
                "text": msg_text,
                "parse_mode": "HTML"
            }

            req = urllib.request.Request(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"}
            )
            urllib.request.urlopen(req)

        except Exception as e:
            logger.error(f"Failed to create task and notify: {e}")

    def complete_task(self, task_id: int):
        task_text = ""
        assignee = ""
        if os.path.exists(TASKS_DB_PATH):
            try:
                conn = sqlite3.connect(TASKS_DB_PATH)
                conn.row_factory = sqlite3.Row
                c = conn.cursor()
                now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
                c.execute("SELECT task_text, assignee FROM tasks WHERE id = ?", (task_id,))
                row = c.fetchone()
                if row:
                    task_text = row["task_text"]
                    assignee = row["assignee"]
                c.execute("UPDATE tasks SET status = 'Done', completed_at = ? WHERE id = ?", (now_str, task_id))
                conn.commit()
                conn.close()
            except Exception as e:
                logger.error(f"Failed to complete task in sqlite: {e}")

        # Update status in Google Sheets & invalidate cache
        creds_json_env = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if creds_json_env:
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                s_clean = creds_json_env.strip().strip("'").strip('"')
                info = json.loads(s_clean)
                if isinstance(info.get("private_key"), str):
                    info["private_key"] = info["private_key"].replace("\\n", "\n")
                creds = Credentials.from_service_account_info(info, scopes=scopes)
                client = gspread.authorize(creds)
                spreadsheet = client.open_by_key("14lJVvDmK9LOAERAo9twp3Ak-FEdvlrzu-8FywP2dTn4")
                sheet = spreadsheet.sheet1
                headers = [str(h).strip() for h in sheet.row_values(1)]
                id_col_vals = sheet.col_values(1)
                target_row = None
                str_id = str(task_id).strip()
                for idx, val in enumerate(id_col_vals):
                    if str(val).strip() == str_id:
                        target_row = idx + 1
                        break
                if target_row:
                    stat_col = headers.index("Статус") + 1 if "Статус" in headers else 7
                    sheet.update_cell(target_row, stat_col, "Done")
                    if not task_text:
                        row_vals = sheet.row_values(target_row)
                        text_idx = headers.index("Текст задачи") if "Текст задачи" in headers else 1
                        ass_idx = headers.index("Исполнитель") if "Исполнитель" in headers else 2
                        task_text = row_vals[text_idx] if len(row_vals) > text_idx else ""
                        assignee = row_vals[ass_idx] if len(row_vals) > ass_idx else ""
                TASKS_SHEETS_CACHE["timestamp"] = 0
            except Exception as e:
                logger.error(f"Failed to complete task in Google Sheets: {e}")

        # Send completion notification to Telegram group
        try:
            bot_token = "8666306951:AAEJ9z2F0t4I2mj2IMPE8TygL6a2k_5ob6g"
            chat_id = "-1002638798110"
            safe_task = (task_text or 'Задача').replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            safe_asgn = (assignee or 'Команда').replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            msg_text = (
                f"✅ <b>ЗАДАЧА #{task_id} ЗАВЕРШЕНА!</b>\n\n"
                f"📌 <b>Задача:</b> {safe_task}\n"
                f"👤 <b>Исполнитель:</b> {safe_asgn}\n\n"
                f"⏳ <i>Ожидает оценки руководителя в приложении...</i>"
            )
            payload = {"chat_id": chat_id, "text": msg_text, "parse_mode": "HTML"}
            req = urllib.request.Request(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"}
            )
            urllib.request.urlopen(req)
        except Exception as e:
            logger.error(f"Failed to send completion notification to Telegram: {e}")

    def rate_task(self, task_id: int, rating: int, rating_comment: str = ""):
        task_text = ""
        assignee = ""
        prev_init_rating = 0
        was_disputed = False

        creds_json_env = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if creds_json_env:
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                s_clean = creds_json_env.strip().strip("'").strip('"')
                info = json.loads(s_clean)
                if isinstance(info.get("private_key"), str):
                    info["private_key"] = info["private_key"].replace("\\n", "\n")
                creds = Credentials.from_service_account_info(info, scopes=scopes)
                client = gspread.authorize(creds)
                spreadsheet = client.open_by_key("14lJVvDmK9LOAERAo9twp3Ak-FEdvlrzu-8FywP2dTn4")
                sheet = spreadsheet.sheet1
                headers = [str(h).strip() for h in sheet.row_values(1)]
                id_col_vals = sheet.col_values(1)
                target_row = None
                str_id = str(task_id).strip()
                for idx, val in enumerate(id_col_vals):
                    if str(val).strip() == str_id:
                        target_row = idx + 1
                        break
                if target_row:
                    row_vals = sheet.row_values(target_row)
                    text_idx = headers.index("Текст задачи") if "Текст задачи" in headers else 1
                    ass_idx = headers.index("Исполнитель") if "Исполнитель" in headers else 2
                    init_rat_idx = headers.index("Первоначальная оценка") if "Первоначальная оценка" in headers else (headers.index("Оценка") if "Оценка" in headers else 7)
                    disp_idx = headers.index("Причина оспаривания") if "Причина оспаривания" in headers else (headers.index("Комментарий / Оспаривание") if "Комментарий / Оспаривание" in headers else 8)

                    task_text = row_vals[text_idx] if len(row_vals) > text_idx else ""
                    assignee = row_vals[ass_idx] if len(row_vals) > ass_idx else ""
                    raw_init = row_vals[init_rat_idx] if len(row_vals) > init_rat_idx else "0"
                    raw_disp = row_vals[disp_idx] if len(row_vals) > disp_idx else ""

                    try:
                        prev_init_rating = int(str(raw_init).replace("/5", "").strip())
                    except Exception:
                        prev_init_rating = 0

                    if raw_disp.strip():
                        was_disputed = True

                    init_col = headers.index("Первоначальная оценка") + 1 if "Первоначальная оценка" in headers else 8
                    final_col = headers.index("Итоговая оценка не меняется") + 1 if "Итоговая оценка не меняется" in headers else (headers.index("Последняя оценка") + 1 if "Последняя оценка" in headers else 10)

                    if not raw_init.strip() or raw_init.strip() == "0":
                        sheet.update_cell(target_row, init_col, f"{rating}/5")
                        prev_init_rating = rating
                    sheet.update_cell(target_row, final_col, f"{rating}/5")

                TASKS_SHEETS_CACHE["timestamp"] = 0
            except Exception as e:
                logger.error(f"Failed to update task rating in Google Sheets: {e}")

        # Send Telegram notification
        try:
            bot_token = os.getenv("BOT_TOKEN", "8951006941:AAH2Wc2j2AH1aCvui1Bflr7puDStzHtwNNI").strip()
            chat_id = "-1002638798110"
            stars_str = "⭐️" * rating

            safe_task = (task_text or 'Задача').replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            safe_asgn = (assignee or 'Команда').replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            if not safe_asgn.startswith("@") and not safe_asgn.startswith("<b>"):
                tag_asgn = f"<b>{safe_asgn}</b>"
            else:
                tag_asgn = safe_asgn

            if was_disputed:
                if rating > prev_init_rating and prev_init_rating > 0:
                    old_stars = "⭐️" * prev_init_rating
                    msg_text = (
                        f"⚖️ <b>РЕЗУЛЬТАТ ОСПАРИВАНИЯ ЗАДАЧИ #{task_id}</b>\n\n"
                        f"🎉 {tag_asgn}, <b>вы выиграли спор! Ваша оценка повышена!</b>\n"
                        f"📌 <b>Задача:</b> {safe_task}\n"
                        f"📉 <b>Первоначальная оценка:</b> {old_stars} ({prev_init_rating}/5)\n"
                        f"📈 <b>Новая (финальная) оценка:</b> {stars_str} ({rating}/5)"
                    )
                else:
                    msg_text = (
                        f"⚖️ <b>РЕЗУЛЬТАТ ОСПАРИВАНИЯ ЗАДАЧИ #{task_id}</b>\n\n"
                        f"❌ {tag_asgn}, <b>вы не выиграли спор. Оценка оставлена без изменений.</b>\n"
                        f"📌 <b>Задача:</b> {safe_task}\n"
                        f"👑 <b>Итоговая оценка:</b> {stars_str} ({rating}/5)"
                    )
                keyboard = None
            else:
                if rating >= 5:
                    msg_text = (
                        f"⭐️ <b>ОЦЕНКА ЗАДАЧИ #{task_id}</b>\n\n"
                        f"📌 <b>Задача:</b> {safe_task}\n"
                        f"👤 <b>Исполнитель:</b> {tag_asgn}\n"
                        f"👑 <b>Оценка руководителя:</b> {stars_str} ({rating}/5)\n\n"
                        f"🎉 <i>Отличная работа! Высокая оценка (5/5) без возможности оспаривания.</i>"
                    )
                    keyboard = None
                else:
                    msg_text = (
                        f"⭐️ <b>ОЦЕНКА ЗАДАЧИ #{task_id}</b>\n\n"
                        f"📌 <b>Задача:</b> {safe_task}\n"
                        f"👤 <b>Исполнитель:</b> {tag_asgn}\n"
                        f"👑 <b>Оценка руководителя:</b> {stars_str} ({rating}/5)\n\n"
                        f"⚖️ <i>Исполнитель {tag_asgn} может оспорить эту оценку, если не согласен:</i>"
                    )
                    keyboard = {
                        "inline_keyboard": [
                            [{"text": "⚖️ Оспорить оценку", "callback_data": f"dispute_task_{task_id}"}]
                        ]
                    }

            payload = {
                "chat_id": chat_id,
                "text": msg_text,
                "parse_mode": "HTML",
                "reply_markup": keyboard
            }

            req = urllib.request.Request(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"}
            )
            urllib.request.urlopen(req)
        except Exception as e:
            logger.error(f"Failed to send rating notification to Telegram: {e}")


    def get_tasks_dynamics(self, date_from=None, date_to=None, assignee_filter=None):
        """Return per-day task dynamics for the selected period and assignee."""
        if not os.path.exists(TASKS_DB_PATH):
            return {"dates": [], "series": []}
        try:
            import datetime as dt

            # Default: last 30 days
            today = dt.date.today()
            if date_to:
                try:
                    end_d = dt.datetime.strptime(date_to, "%Y-%m-%d").date()
                except Exception:
                    end_d = today
            else:
                end_d = today

            if date_from:
                try:
                    start_d = dt.datetime.strptime(date_from, "%Y-%m-%d").date()
                except Exception:
                    start_d = end_d - dt.timedelta(days=29)
            else:
                start_d = end_d - dt.timedelta(days=29)

            # Build all dates in range
            all_dates = []
            cur = start_d
            while cur <= end_d:
                all_dates.append(cur)
                cur += dt.timedelta(days=1)

            conn = sqlite3.connect(TASKS_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # Get all tasks
            c.execute("SELECT assignee, created_at, completed_at, status, rating FROM tasks")
            rows = c.fetchall()
            conn.close()

            # Parse and filter tasks
            TEAM_LEADS = ["@isslamov", "@axi0603", "@Silent_trickster"]

            def parse_date(s):
                if not s:
                    return None
                for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                    try:
                        return dt.datetime.strptime(s.strip(), fmt).date()
                    except Exception:
                        pass
                return None

            # Build per-lead, per-day counters
            leads = TEAM_LEADS if not assignee_filter or assignee_filter == "all" else [assignee_filter]

            date_strs = [d.strftime("%d.%m") for d in all_dates]

            series = []
            for lead in leads:
                created_by_day = {}
                done_by_day = {}
                rating_by_day = {}

                for row in rows:
                    if row["assignee"] != lead:
                        continue

                    c_date = parse_date(row["created_at"])
                    if c_date and start_d <= c_date <= end_d:
                        ds = c_date.strftime("%d.%m")
                        created_by_day[ds] = created_by_day.get(ds, 0) + 1

                    if row["status"] == "Done" and row["completed_at"]:
                        d_date = parse_date(row["completed_at"])
                        if d_date and start_d <= d_date <= end_d:
                            ds = d_date.strftime("%d.%m")
                            done_by_day[ds] = done_by_day.get(ds, 0) + 1

                    if row["rating"] and row["rating"] > 0:
                        c_date2 = parse_date(row["created_at"])
                        if c_date2 and start_d <= c_date2 <= end_d:
                            ds = c_date2.strftime("%d.%m")
                            if ds not in rating_by_day:
                                rating_by_day[ds] = []
                            rating_by_day[ds].append(row["rating"])

                # Build arrays in order
                created_arr = [created_by_day.get(ds, 0) for ds in date_strs]
                done_arr = [done_by_day.get(ds, 0) for ds in date_strs]
                avg_rating_arr = [
                    round(sum(rating_by_day[ds]) / len(rating_by_day[ds]), 1) if rating_by_day.get(ds) else None
                    for ds in date_strs
                ]

                short_name = lead.replace("@", "")
                series.append({
                    "lead": lead,
                    "short_name": short_name,
                    "created": created_arr,
                    "done": done_arr,
                    "avg_rating": avg_rating_arr,
                })

            # Totals across all leads (for "All" view)
            all_created = [sum(s["created"][i] for s in series) for i in range(len(date_strs))]
            all_done = [sum(s["done"][i] for s in series) for i in range(len(date_strs))]

            return {
                "dates": date_strs,
                "date_range": {
                    "from": start_d.strftime("%d.%m.%Y"),
                    "to": end_d.strftime("%d.%m.%Y"),
                    "days": len(all_dates)
                },
                "series": series,
                "totals": {
                    "created": all_created,
                    "done": all_done,
                }
            }
        except Exception as e:
            logger.error(f"Failed to get tasks dynamics: {e}")
            return {"dates": [], "series": [], "totals": {"created": [], "done": []}}

    def archive_current_period(self, period_label: str, period_start: str, period_end: str):
        """Archive current period stats per assignee into task_archive table."""
        if not os.path.exists(TASKS_DB_PATH):
            return {"error": "tasks.db not found"}
        try:
            conn = sqlite3.connect(TASKS_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # Ensure table exists
            c.execute("""
                CREATE TABLE IF NOT EXISTS task_archive (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    period_label TEXT NOT NULL,
                    period_start TEXT NOT NULL,
                    period_end TEXT NOT NULL,
                    assignee TEXT NOT NULL,
                    total_tasks INTEGER DEFAULT 0,
                    done_tasks INTEGER DEFAULT 0,
                    avg_rating REAL DEFAULT 0,
                    efficiency INTEGER DEFAULT 0,
                    archived_at TEXT
                )
            """)

            now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

            # Get stats per assignee for all tasks
            c.execute("""
                SELECT assignee,
                       COUNT(*) as total,
                       SUM(CASE WHEN status='Done' THEN 1 ELSE 0 END) as done,
                       AVG(CASE WHEN rating > 0 THEN rating ELSE NULL END) as avg_r
                FROM tasks
                GROUP BY assignee
            """)
            rows = c.fetchall()

            for row in rows:
                assignee = row["assignee"] or "Команда"
                total = row["total"] or 0
                done = row["done"] or 0
                avg_r = round(row["avg_r"] or 0, 1)
                eff = round((done / total * 100)) if total > 0 else 0

                c.execute("""
                    INSERT INTO task_archive
                    (period_label, period_start, period_end, assignee, total_tasks, done_tasks, avg_rating, efficiency, archived_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (period_label, period_start, period_end, assignee, total, done, avg_r, eff, now_str))

            conn.commit()
            conn.close()
            return {"status": "ok", "archived": len(rows), "period": period_label}
        except Exception as e:
            logger.error(f"Failed to archive period: {e}")
            return {"error": str(e)}

    def get_archives(self):
        """Get all archived periods with period-over-period comparison."""
        if not os.path.exists(TASKS_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(TASKS_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            c.execute("""
                CREATE TABLE IF NOT EXISTS task_archive (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    period_label TEXT NOT NULL,
                    period_start TEXT NOT NULL,
                    period_end TEXT NOT NULL,
                    assignee TEXT NOT NULL,
                    total_tasks INTEGER DEFAULT 0,
                    done_tasks INTEGER DEFAULT 0,
                    avg_rating REAL DEFAULT 0,
                    efficiency INTEGER DEFAULT 0,
                    archived_at TEXT
                )
            """)

            # Get all periods
            c.execute("SELECT DISTINCT period_label, period_start, period_end, archived_at FROM task_archive ORDER BY archived_at DESC")
            periods = [dict(r) for r in c.fetchall()]

            result = []
            for p in periods:
                label = p["period_label"]
                c.execute("""
                    SELECT assignee, total_tasks, done_tasks, avg_rating, efficiency
                    FROM task_archive WHERE period_label = ?
                    ORDER BY efficiency DESC
                """, (label,))
                leads = [dict(r) for r in c.fetchall()]
                result.append({
                    "period_label": label,
                    "period_start": p["period_start"],
                    "period_end": p["period_end"],
                    "archived_at": p["archived_at"],
                    "leads": leads
                })

            conn.close()
            return result
        except Exception as e:
            logger.error(f"Failed to get archives: {e}")
            return []

    def get_team_leads_task_stats(self, month=None):
        date_pattern1 = None
        date_pattern2 = None
        if month and len(month.split("-")) == 2:
            yyyy, mm = month.split("-")
            date_pattern1 = f"%.{mm}.{yyyy}%"
            date_pattern2 = f"{yyyy}-{mm}%"

        team_leads = [
            {"name": "Ильясбек (@isslamov)", "role": "Тимлид", "patterns": ["%isslaamov%", "%isslamov%", "%Ильясбек%", "%ilyas%"]},
            {"name": "Мужахидбек (@axi0603)", "role": "Тимлид", "patterns": ["%axi0603%", "%axi%", "%мужахид%", "%mujahid%"]},
            {"name": "Жахабек (@Silent_trickster)", "role": "Тимлид", "patterns": ["%Silent_trickster%", "%silenttrickster%", "%jaxa%", "%жаха%", "%jakha%"]}
        ]
        results = []
        if os.getenv("DATABASE_URL") or os.path.exists(TASKS_DB_PATH):
            try:
                conn = sqlite3.connect(TASKS_DB_PATH)
                c = conn.cursor()
                for tl in team_leads:
                    conds_base = " OR ".join(["assignee LIKE ?" for _ in tl["patterns"]])
                    params_base = list(tl["patterns"])

                    if date_pattern1 and date_pattern2:
                        query_tot = f"SELECT COUNT(*) FROM tasks WHERE ({conds_base}) AND (created_at LIKE ? OR created_at LIKE ?)"
                        c.execute(query_tot, params_base + [date_pattern1, date_pattern2])
                    else:
                        query_tot = f"SELECT COUNT(*) FROM tasks WHERE ({conds_base})"
                        c.execute(query_tot, params_base)
                    total = c.fetchone()[0]

                    if date_pattern1 and date_pattern2:
                        query_act = f"SELECT COUNT(*) FROM tasks WHERE ({conds_base}) AND status != 'Done' AND (created_at LIKE ? OR created_at LIKE ?)"
                        c.execute(query_act, params_base + [date_pattern1, date_pattern2])
                    else:
                        query_act = f"SELECT COUNT(*) FROM tasks WHERE ({conds_base}) AND status != 'Done'"
                        c.execute(query_act, params_base)
                    active = c.fetchone()[0]

                    if date_pattern1 and date_pattern2:
                        query_done = f"SELECT COUNT(*) FROM tasks WHERE ({conds_base}) AND status = 'Done' AND (created_at LIKE ? OR created_at LIKE ?)"
                        c.execute(query_done, params_base + [date_pattern1, date_pattern2])
                    else:
                        query_done = f"SELECT COUNT(*) FROM tasks WHERE ({conds_base}) AND status = 'Done'"
                        c.execute(query_done, params_base)
                    done = c.fetchone()[0]

                    if date_pattern1 and date_pattern2:
                        query_rate = f"SELECT AVG(COALESCE(rating, 5)) FROM tasks WHERE ({conds_base}) AND status = 'Done' AND rating > 0 AND (created_at LIKE ? OR created_at LIKE ?)"
                        c.execute(query_rate, params_base + [date_pattern1, date_pattern2])
                    else:
                        query_rate = f"SELECT AVG(COALESCE(rating, 5)) FROM tasks WHERE ({conds_base}) AND status = 'Done' AND rating > 0"
                        c.execute(query_rate, params_base)
                    avg_r = c.fetchone()[0]
                    avg_rating = round(float(avg_r), 1) if (avg_r is not None and done > 0) else 0.0

                    percent = round((done / total * 100)) if total > 0 else 0
                    results.append({
                        "name": tl["name"],
                        "role": tl["role"],
                        "total": total,
                        "active": active,
                        "done": done,
                        "percent": percent,
                        "avg_rating": avg_rating
                    })
                conn.close()
                return results
            except Exception as e:
                logger.error(f"Failed to get team leads task stats: {e}")

        return [
            {"name": "Ильясбек (@isslamov)", "role": "Тимлид", "total": 0, "active": 0, "done": 0, "percent": 0, "avg_rating": 0.0},
            {"name": "Мужахидбек (@axi0603)", "role": "Тимлид", "total": 0, "active": 0, "done": 0, "percent": 0, "avg_rating": 0.0},
            {"name": "Жахабек (@Silent_trickster)", "role": "Тимлид", "total": 0, "active": 0, "done": 0, "percent": 0, "avg_rating": 0.0}
        ]

    def get_users_data(self):
        if not os.path.exists(BIKES_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT user_id, username, full_name, role, is_active FROM users ORDER BY rowid ASC")
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return rows
        except Exception as e:
            logger.error(f"Failed to get users: {e}")
            return []

    def toggle_user_access(self, user_id: int, is_active: int):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            c.execute("UPDATE users SET is_active = ? WHERE user_id = ?", (is_active, user_id))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to toggle user access: {e}")

    def get_rich_cities(self):
        if not os.path.exists(BIKES_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT id, name, total_bikes FROM rich_cities ORDER BY id ASC")
            cities = [dict(r) for r in c.fetchall()]

            for city in cities:
                clean_cname = city['name'].replace('(Rich)', '').strip()
                c.execute("SELECT issued FROM rich_reports WHERE city LIKE ? OR city = ? ORDER BY id DESC LIMIT 1", (f"%{clean_cname}%", city['name']))
                rep = c.fetchone()
                issued = int(rep['issued']) if rep and rep['issued'] is not None and str(rep['issued']).isdigit() else 0
                pct = round((issued / city['total_bikes']) * 100) if city['total_bikes'] > 0 and issued > 0 else 0
                city['issued'] = issued
                city['percent_online'] = pct

            conn.close()
            return cities
        except Exception as e:
            logger.error(f"Failed to get rich cities: {e}")
            return []

    def add_rich_city(self, name: str, total_bikes: int):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            now_str = datetime.datetime.now().strftime("%d.%m.%Y")
            c.execute("INSERT OR IGNORE INTO rich_cities (name, total_bikes, created_at) VALUES (?, ?, ?)", (name, total_bikes, now_str))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to add rich city: {e}")

    def update_rich_city(self, city_id: int, total_bikes: int):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            c.execute("UPDATE rich_cities SET total_bikes = ? WHERE id = ?", (total_bikes, city_id))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to update rich city: {e}")

    def get_rich_reports(self):
        if not os.path.exists(BIKES_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT id, username, city, report_date, issued, returned, comment FROM rich_reports ORDER BY id DESC LIMIT 20")
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return rows
        except Exception as e:
            logger.error(f"Failed to get rich reports: {e}")
            return []

    def get_rich_stats(self):
        tot_rich_fleet = 0
        active_rich_bots = 0
        if os.path.exists(BIKES_DB_PATH):
            try:
                conn = sqlite3.connect(BIKES_DB_PATH)
                conn.row_factory = sqlite3.Row
                c = conn.cursor()
                c.execute("SELECT SUM(total_bikes) as tot FROM rich_cities")
                row = c.fetchone()
                tot_rich_fleet = row["tot"] if row and row["tot"] else 0

                c.execute("SELECT COUNT(*) as cnt FROM managed_bots WHERE project_type = 'Rich' AND is_active = 1")
                b_row = c.fetchone()
                active_rich_bots = b_row["cnt"] if b_row else 0

                conn.close()
            except Exception:
                pass
        return {
            "status": "Active",
            "uptime": "100%",
            "total_rich_fleet": tot_rich_fleet,
            "active_rich_bots": active_rich_bots,
        }

    def get_payroll_data(self):
        if not os.path.exists(BIKES_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT id, employee_name, advance_amount, salary_amount, tax_amount, total_fot, payment_date FROM payroll ORDER BY id DESC LIMIT 20")
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return rows
        except Exception as e:
            logger.error(f"Failed to get payroll data: {e}")
            return []

    def add_payroll_entry(self, emp_name: str, advance: float, salary: float, tax: float, fot: float):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
            c.execute(
                "INSERT INTO payroll (employee_name, advance_amount, salary_amount, tax_amount, total_fot, payment_date) VALUES (?, ?, ?, ?, ?, ?)",
                (emp_name, advance, salary, tax, fot, now_str)
            )
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to add payroll entry: {e}")

    def get_bots_data(self):
        if not os.path.exists(BIKES_DB_PATH):
            return []
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT id, bot_name, bot_token, bot_username, project_type, city_name, report_type, is_active FROM managed_bots ORDER BY id ASC")
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return rows
        except Exception as e:
            logger.error(f"Failed to get bots data: {e}")
            return []

    def add_managed_bot(self, bot_name: str, bot_token: str, bot_username: str, project_type: str, city_name: str, report_type: str):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            now_str = datetime.datetime.now().strftime("%d.%m.%Y")
            c.execute(
                "INSERT INTO managed_bots (bot_name, bot_token, bot_username, project_type, city_name, report_type, is_active, created_at) VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
                (bot_name, bot_token, bot_username, project_type, city_name, report_type, now_str)
            )
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to add managed bot: {e}")

    def toggle_managed_bot(self, bot_id: int, is_active: int):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            c.execute("UPDATE managed_bots SET is_active = ? WHERE id = ?", (is_active, bot_id))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to toggle managed bot: {e}")

    def update_managed_bot(self, bot_id: int, bot_name: str, bot_token: str, bot_username: str, project_type: str, city_name: str, report_type: str):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            c.execute(
                "UPDATE managed_bots SET bot_name = ?, bot_token = ?, bot_username = ?, project_type = ?, city_name = ?, report_type = ? WHERE id = ?",
                (bot_name, bot_token, bot_username, project_type, city_name, report_type, bot_id)
            )
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to update managed bot: {e}")

    def delete_managed_bot(self, bot_id: int):
        if not os.path.exists(BIKES_DB_PATH):
            return
        try:
            conn = sqlite3.connect(BIKES_DB_PATH)
            c = conn.cursor()
            c.execute("DELETE FROM managed_bots WHERE id = ?", (bot_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Failed to delete managed bot: {e}")

def init_local_master_dbs():
    try:
        conn = sqlite3.connect(BIKES_DB_PATH)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS cities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                has_bike_types INTEGER DEFAULT 0,
                total_bikes INTEGER DEFAULT 80,
                created_at TEXT NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS bike_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                city TEXT DEFAULT '',
                report_date TEXT NOT NULL,
                issued TEXT NOT NULL,
                returned TEXT NOT NULL,
                total_in_trip TEXT NOT NULL,
                new_bikes TEXT NOT NULL,
                old_bikes TEXT NOT NULL,
                broken_bikes TEXT NOT NULL,
                return_reasons TEXT NOT NULL,
                comment TEXT,
                created_at TEXT NOT NULL
            )
        """)
        default_cities = [
            ("Ташкент", 1670, 0),
            ("Самарканд", 200, 0),
            ("Фергана", 80, 0),
            ("Андижан", 50, 0),
            ("Бухара", 30, 0),
            ("Навои", 30, 0),
            ("Карши", 30, 0),
            ("Ургенч", 30, 0),
            ("Нукус", 30, 0),
            ("Коканд", 25, 0),
            ("Наманган", 25, 0),
        ]
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for c_name, c_bikes, c_types in default_cities:
            c.execute("SELECT id FROM cities WHERE name = ?", (c_name,))
            r = c.fetchone()
            if r:
                c.execute("UPDATE cities SET total_bikes = ?, has_bike_types = ? WHERE id = ?", (c_bikes, c_types, r[0]))
            else:
                c.execute("INSERT INTO cities (name, has_bike_types, total_bikes, created_at) VALUES (?, ?, ?, ?)", (c_name, c_types, c_bikes, now))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Error init local bikes db: {e}")

    try:
        conn_t = sqlite3.connect(TASKS_DB_PATH)
        ct = conn_t.cursor()
        ct.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_text TEXT NOT NULL,
                assignee TEXT,
                author TEXT,
                sla_deadline TEXT,
                created_at TEXT,
                status TEXT DEFAULT 'Active',
                priority TEXT DEFAULT 'Medium',
                city TEXT DEFAULT 'Ташкент',
                rating INTEGER DEFAULT 0,
                rating_comment TEXT
            )
        """)
        conn_t.commit()
        conn_t.close()
    except Exception as e:
        logger.error(f"Error init local tasks db: {e}")

def run_master_server(port=8085):
    init_local_master_dbs()
    port = int(os.getenv("PORT", str(port)))
    ThreadingHTTPServer.allow_reuse_address = True
    server = ThreadingHTTPServer(("0.0.0.0", port), MasterHubHandler)
    logger.info(f"Master Hub Server running on 0.0.0.0:{port}.")
    server.serve_forever()

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    run_master_server(8085)
